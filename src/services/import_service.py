"""Excel SQL知识库导入服务。"""

from __future__ import annotations

import hashlib
import re
from dataclasses import (
    asdict,
    dataclass,
)
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select

from src.auth.auth_service import (
    AuthenticatedUser,
    require_role,
)
from src.db.database import session_scope
from src.db.models import (
    ImportBatch,
    SQLRecord,
    utc_now,
)
from src.services.knowledge_service import (
    KnowledgeService,
    SQLRecordData,
)


@dataclass(frozen=True)
class ParsedSQLRow:
    row_number: int
    sheet_name: str
    record_id: int | None
    data: SQLRecordData


@dataclass(frozen=True)
class ImportCandidate:
    row_number: int
    sheet_name: str
    data: SQLRecordData
    existing_record_id: int | None = None
    existing_version: int | None = None


@dataclass
class ImportPreview:
    file_path: Path
    file_name: str
    file_hash: str
    mode: str
    inserts: list[ImportCandidate]
    updates: list[ImportCandidate]
    unchanged: list[ImportCandidate]
    invalid: list[dict[str, str]]

    @property
    def total_count(self) -> int:
        return (
            len(self.inserts)
            + len(self.updates)
            + len(self.unchanged)
            + len(self.invalid)
        )


def _clean_value(value) -> str:
    if value is None:
        return ""

    if (
        isinstance(value, float)
        and value.is_integer()
    ):
        return str(int(value))

    return str(value).strip()


def _normalize_header(value) -> str:
    return re.sub(
        r"[\s_\-—:：()（）/\\【】\[\]]+",
        "",
        _clean_value(value).lower(),
    )


HEADER_ALIASES = {
    "record_id": {
        "记录id",
        "数据库记录id",
        "内部记录id",
        "记录id勿修改",
    },
    "sql_id": {
        "sqlid",
        "sql编号",
        "sql序号",
    },
    "source_sql_id": {
        "原文件sqlid",
        "来源sqlid",
        "原始sqlid",
    },
    "source_row": {
        "原文件行号",
        "来源行号",
        "原始行号",
    },
    "business_domain": {
        "业务域",
        "业务分类",
    },
    "function_theme": {
        "功能主题",
        "功能名称",
        "sql功能",
        "查询功能",
    },
    "step": {
        "步骤",
        "执行步骤",
    },
    "function_type": {
        "功能类型",
        "功能分类",
    },
    "statement_type": {
        "语句类型",
        "sql类型",
    },
    "source_tables": {
        "涉及来源表",
        "来源表",
        "涉及表",
    },
    "table_completeness": {
        "表资料完整性",
        "表完整性",
    },
    "field_completeness": {
        "字段资料完整性",
        "字段完整性",
    },
    "missing_tables": {
        "缺失占位表",
        "缺失表",
    },
    "unregistered_fields": {
        "未登记字段",
    },
    "field_ownership_unknown": {
        "字段归属不明",
    },
    "unexplained_fields": {
        "未解释字段",
        "缺少释义字段",
    },
    "sql_text": {
        "sql正文",
        "sql语句",
        "sql详细语句",
        "标准sql",
        "hivesql",
        "sql内容",
    },
    "notes": {
        "备注",
        "原始标题注释",
    },
    "source_file": {
        "来源文件",
        "原文件名",
    },
    "source_sheet": {
        "来源工作表",
        "原工作表",
    },
}


def _find_column(
    headers: list[str],
    field_name: str,
) -> int | None:
    aliases = HEADER_ALIASES[
        field_name
    ]

    for index, header in enumerate(headers):
        if header in aliases:
            return index

    if field_name == "function_theme":
        for index, header in enumerate(headers):
            if (
                "功能" in header
                and "类型" not in header
                and "分类" not in header
            ):
                return index

    if field_name == "sql_text":
        for index, header in enumerate(headers):
            if (
                "sql" in header
                and "id" not in header
                and (
                    "正文" in header
                    or "语句" in header
                    or "内容" in header
                )
            ):
                return index

    return None


def _find_header_row(
    rows: list[tuple],
) -> tuple[
    int,
    dict[str, int | None],
] | None:
    for row_index, row in enumerate(rows[:50]):
        headers = [
            _normalize_header(cell)
            for cell in row
        ]

        column_map = {
            field: _find_column(
                headers,
                field,
            )
            for field in HEADER_ALIASES
        }

        if (
            column_map["function_theme"]
            is not None
            and column_map["sql_text"]
            is not None
        ):
            return row_index, column_map

    return None


def _file_hash(path: Path) -> str:
    digest = hashlib.sha256()

    with path.open("rb") as file:
        for block in iter(
            lambda: file.read(1024 * 1024),
            b"",
        ):
            digest.update(block)

    return digest.hexdigest()


def _parse_optional_int(
    value: str,
) -> int | None:
    if not value:
        return None

    try:
        return int(value)

    except ValueError:
        return None


def parse_sql_xlsx(
    path: str | Path,
) -> tuple[
    list[ParsedSQLRow],
    list[dict[str, str]],
]:
    file_path = Path(path)

    if not file_path.exists():
        raise FileNotFoundError(
            f"文件不存在：{file_path}"
        )

    workbook = load_workbook(
        file_path,
        read_only=True,
        data_only=True,
    )

    parsed_rows = []
    invalid_rows = []
    seen_record_ids: set[int] = set()

    for sheet in workbook.worksheets:
        rows = list(
            sheet.iter_rows(
                values_only=True
            )
        )

        if not rows:
            continue

        header_result = _find_header_row(rows)

        if header_result is None:
            continue

        header_index, column_map = (
            header_result
        )

        for row_number, row in enumerate(
            rows[header_index + 1:],
            start=header_index + 2,
        ):
            values = [
                _clean_value(cell)
                for cell in row
            ]

            def value_of(
                field_name: str,
            ) -> str:
                index = column_map.get(
                    field_name
                )

                if (
                    index is None
                    or index >= len(values)
                ):
                    return ""

                return values[index]

            function_theme = value_of(
                "function_theme"
            )
            sql_text = value_of("sql_text")

            if (
                not function_theme
                and not sql_text
            ):
                continue

            missing = []

            if not function_theme:
                missing.append("功能主题")

            if not sql_text:
                missing.append("SQL正文")

            if missing:
                invalid_rows.append(
                    {
                        "sheet": sheet.title,
                        "row": str(row_number),
                        "sql_id": value_of(
                            "sql_id"
                        ),
                        "reason": (
                            "缺少必填字段："
                            + "、".join(missing)
                        ),
                    }
                )
                continue

            raw_record_id = value_of(
                "record_id"
            )
            record_id = _parse_optional_int(
                raw_record_id
            )

            if raw_record_id and record_id is None:
                invalid_rows.append(
                    {
                        "sheet": sheet.title,
                        "row": str(row_number),
                        "sql_id": value_of(
                            "sql_id"
                        ),
                        "reason": (
                            "记录ID必须是整数。"
                        ),
                    }
                )
                continue

            if record_id is not None:
                if record_id in seen_record_ids:
                    invalid_rows.append(
                        {
                            "sheet": sheet.title,
                            "row": str(row_number),
                            "sql_id": value_of(
                                "sql_id"
                            ),
                            "reason": (
                                "文件中记录ID重复。"
                            ),
                        }
                    )
                    continue

                seen_record_ids.add(record_id)

            explicit_source_sql_id = value_of(
                "source_sql_id"
            )

            displayed_sql_id = value_of(
                "sql_id"
            )

            # 新外部文件的SQL_ID属于来源编号
            # 系统导出文件优先保留原文件SQL_ID
            source_sql_id = (
                explicit_source_sql_id
                or (
                    displayed_sql_id
                    if record_id is None
                    else None
                )
                or None
            )

            explicit_source_row = (
                _parse_optional_int(
                    value_of("source_row")
                )
            )

            data = SQLRecordData(
                function_theme=function_theme,
                sql_text=sql_text,
                source_sql_id=source_sql_id,
                source_row=(
                    explicit_source_row
                    or row_number
                ),
                business_domain=(
                    value_of(
                        "business_domain"
                    )
                    or None
                ),
                step=(
                    value_of("step")
                    or None
                ),
                function_type=(
                    value_of(
                        "function_type"
                    )
                    or None
                ),
                statement_type=(
                    value_of(
                        "statement_type"
                    )
                    or None
                ),
                source_tables=(
                    value_of(
                        "source_tables"
                    )
                    or None
                ),
                table_completeness=(
                    value_of(
                        "table_completeness"
                    )
                    or None
                ),
                field_completeness=(
                    value_of(
                        "field_completeness"
                    )
                    or None
                ),
                missing_tables=(
                    value_of(
                        "missing_tables"
                    )
                    or None
                ),
                unregistered_fields=(
                    value_of(
                        "unregistered_fields"
                    )
                    or None
                ),
                field_ownership_unknown=(
                    value_of(
                        "field_ownership_unknown"
                    )
                    or None
                ),
                unexplained_fields=(
                    value_of(
                        "unexplained_fields"
                    )
                    or None
                ),
                notes=(
                    value_of("notes")
                    or None
                ),
                source_file=(
                    value_of("source_file")
                    or file_path.name
                ),
                source_sheet=(
                    value_of("source_sheet")
                    or sheet.title
                ),
            )

            parsed_rows.append(
                ParsedSQLRow(
                    row_number=row_number,
                    sheet_name=sheet.title,
                    record_id=record_id,
                    data=data,
                )
            )

    workbook.close()

    if not parsed_rows and not invalid_rows:
        raise ValueError(
            "没有识别到SQL明细记录。"
        )

    return parsed_rows, invalid_rows


def _data_signature(
    data: SQLRecordData,
) -> dict:
    result = asdict(data)

    for key, value in result.items():
        if isinstance(value, str):
            result[key] = value.strip() or None

    return result


def _record_signature(
    record: SQLRecord,
) -> dict:
    return _data_signature(
        SQLRecordData(
            function_theme=record.function_theme,
            sql_text=record.sql_text,
            source_sql_id=record.source_sql_id,
            source_row=record.source_row,
            business_domain=record.business_domain,
            step=record.step,
            function_type=record.function_type,
            statement_type=(
                record.statement_type
            ),
            source_tables=record.source_tables,
            table_completeness=(
                record.table_completeness
            ),
            field_completeness=(
                record.field_completeness
            ),
            missing_tables=record.missing_tables,
            unregistered_fields=(
                record.unregistered_fields
            ),
            field_ownership_unknown=(
                record.field_ownership_unknown
            ),
            unexplained_fields=(
                record.unexplained_fields
            ),
            notes=record.notes,
            source_file=record.source_file,
            source_sheet=record.source_sheet,
        )
    )


class ImportService:
    """Excel导入服务。"""

    def __init__(self) -> None:
        self.knowledge_service = (
            KnowledgeService()
        )

    def preview(
        self,
        path: str | Path,
        current_user: AuthenticatedUser,
    ) -> ImportPreview:
        require_role(
            current_user,
            "editor",
            "admin",
        )

        file_path = Path(path)

        parsed_rows, invalid = parse_sql_xlsx(
            file_path
        )

        contains_record_ids = any(
            row.record_id is not None
            for row in parsed_rows
        )

        mode = (
            "update"
            if contains_record_ids
            else "append"
        )

        file_hash = _file_hash(file_path)

        with session_scope() as session:
            existing_records = list(
                session.scalars(
                    select(SQLRecord)
                ).all()
            )

            # 新增文件完全相同时，禁止重复追加
            if mode == "append":
                duplicate_batch = session.scalar(
                    select(ImportBatch).where(
                        ImportBatch.file_hash
                        == file_hash,
                        ImportBatch.status.in_(
                            ["success", "partial"]
                        ),
                    )
                )

                if duplicate_batch is not None:
                    raise ValueError(
                        "该新增文件已经导入过，"
                        "不能重复追加。"
                    )

        existing_by_id = {
            record.id: record
            for record in existing_records
        }

        inserts = []
        updates = []
        unchanged = []

        for row in parsed_rows:
            if row.record_id is None:
                inserts.append(
                    ImportCandidate(
                        row_number=row.row_number,
                        sheet_name=row.sheet_name,
                        data=row.data,
                    )
                )
                continue

            existing = existing_by_id.get(
                row.record_id
            )

            if existing is None:
                invalid.append(
                    {
                        "sheet": row.sheet_name,
                        "row": str(row.row_number),
                        "sql_id": "",
                        "reason": (
                            f"记录ID {row.record_id} "
                            "在数据库中不存在。"
                        ),
                    }
                )
                continue

            if existing.deleted_at is not None:
                invalid.append(
                    {
                        "sheet": row.sheet_name,
                        "row": str(row.row_number),
                        "sql_id": "",
                        "reason": (
                            f"记录ID {row.record_id} "
                            "已经被删除。"
                        ),
                    }
                )
                continue

            candidate = ImportCandidate(
                row_number=row.row_number,
                sheet_name=row.sheet_name,
                data=row.data,
                existing_record_id=existing.id,
                existing_version=(
                    existing.version
                ),
            )

            if (
                _data_signature(row.data)
                == _record_signature(existing)
            ):
                unchanged.append(candidate)
            else:
                updates.append(candidate)

        return ImportPreview(
            file_path=file_path,
            file_name=file_path.name,
            file_hash=file_hash,
            mode=mode,
            inserts=inserts,
            updates=updates,
            unchanged=unchanged,
            invalid=invalid,
        )

    def apply(
        self,
        preview: ImportPreview,
        current_user: AuthenticatedUser,
    ) -> dict:
        require_role(
            current_user,
            "editor",
            "admin",
        )

        with session_scope() as session:
            batch = ImportBatch(
                file_name=preview.file_name,
                file_hash=preview.file_hash,
                import_mode=preview.mode,
                status="processing",
                unchanged_count=len(
                    preview.unchanged
                ),
                invalid_count=len(
                    preview.invalid
                ),
                uploaded_by=current_user.id,
            )

            session.add(batch)
            session.flush()
            batch_id = batch.id

        inserted = 0
        updated = 0
        failures = []

        # 按Excel中的原始顺序追加，
        # 保证重新分配的SQL_ID顺序正确。
        for candidate in preview.inserts:
            try:
                self.knowledge_service.create_record(
                    candidate.data,
                    current_user,
                )
                inserted += 1

            except Exception as exc:
                failures.append(
                    {
                        "action": "INSERT",
                        "row": candidate.row_number,
                        "error": str(exc),
                    }
                )

        for candidate in preview.updates:
            try:
                self.knowledge_service.update_record(
                    record_id=int(
                        candidate.existing_record_id
                    ),
                    data=candidate.data,
                    expected_version=int(
                        candidate.existing_version
                    ),
                    current_user=current_user,
                )
                updated += 1

            except Exception as exc:
                failures.append(
                    {
                        "action": "UPDATE",
                        "row": candidate.row_number,
                        "error": str(exc),
                    }
                )

        failed_count = len(failures)

        if failed_count == 0:
            status = (
                "partial"
                if preview.invalid
                else "success"
            )
        elif inserted or updated:
            status = "partial"
        else:
            status = "failed"

        with session_scope() as session:
            batch = session.get(
                ImportBatch,
                batch_id,
            )

            batch.status = status
            batch.insert_count = inserted
            batch.update_count = updated
            batch.unchanged_count = len(
                preview.unchanged
            )
            batch.invalid_count = (
                len(preview.invalid)
                + failed_count
            )
            batch.error_message = (
                str(failures)
                if failures
                else None
            )
            batch.completed_at = utc_now()

        return {
            "batch_id": batch_id,
            "status": status,
            "inserted": inserted,
            "updated": updated,
            "unchanged": len(
                preview.unchanged
            ),
            "invalid": len(preview.invalid),
            "failed": failed_count,
            "failed_items": failures,
        }
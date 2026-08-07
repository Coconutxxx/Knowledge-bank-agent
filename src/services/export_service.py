"""从数据库导出最新版SQL知识库。"""

from __future__ import annotations
from copy import copy
import json
from dataclasses import dataclass
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
)
from sqlalchemy import select

from src.auth.auth_service import (
    AuthenticatedUser,
    require_role,
)
from src.db.database import session_scope
from src.db.models import (
    SQLRecord,
    utc_now,
)
from src.db.repositories import (
    AuditLogRepository,
)
from datetime import (
    datetime,
    timezone,
)
from zoneinfo import ZoneInfo
from openpyxl.utils import get_column_letter

@dataclass(frozen=True)
class ExportResult:
    data: bytes
    file_name: str
    record_count: int


EXPORT_HEADERS = [
    "记录ID（勿修改）",
    "SQL_ID",
    "原文件SQL_ID",
    "原文件行号",
    "业务域",
    "功能主题",
    "步骤",
    "功能类型",
    "语句类型",
    "涉及来源表",
    "表资料完整性",
    "字段资料完整性",
    "缺失/占位表",
    "未登记字段",
    "字段归属不明",
    "未解释字段",
    "SQL正文",
    "备注",
    "来源文件",
    "来源工作表",
    "记录版本",
    "创建时间",
    "更新时间",
]

CHINA_TIMEZONE = ZoneInfo(
    "Asia/Shanghai"
)

EXCEL_DATETIME_FORMAT = (
    "yyyy-mm-dd hh:mm:ss"
)

def to_china_excel_datetime(
    value: datetime | str | None,
) -> datetime | None:
    """
    将数据库UTC时间转换为中国时间。

    最终返回不带时区信息的datetime，
    因为Excel不支持带时区的datetime。
    """

    if value is None:
        return None

    if isinstance(value, str):
        cleaned_value = value.strip()

        if not cleaned_value:
            return None

        # 兼容末尾使用Z表示UTC的时间
        cleaned_value = (
            cleaned_value.replace(
                "Z",
                "+00:00",
            )
        )

        value = datetime.fromisoformat(
            cleaned_value
        )

    # SQLite有时会返回不带时区的时间，
    # 当前系统数据库约定这种时间也是UTC
    if value.tzinfo is None:
        value = value.replace(
            tzinfo=timezone.utc
        )

    china_time = value.astimezone(
        CHINA_TIMEZONE
    )

    # Excel不能写入带时区的datetime
    return china_time.replace(
        tzinfo=None,
        microsecond=0,
    )


def current_china_time() -> datetime:
    """
    返回当前中国时间，供导出时间使用。
    """

    china_time = datetime.now(
        timezone.utc
    ).astimezone(
        CHINA_TIMEZONE
    )

    return china_time.replace(
        tzinfo=None,
        microsecond=0,
    )

def _record_to_row(
    record: SQLRecord,
) -> list:
    return [
        record.id,
        record.sql_id,
        record.source_sql_id,
        record.source_row,
        record.business_domain,
        record.function_theme,
        record.step,
        record.function_type,
        record.statement_type,
        record.source_tables,
        record.table_completeness,
        record.field_completeness,
        record.missing_tables,
        record.unregistered_fields,
        record.field_ownership_unknown,
        record.unexplained_fields,
        record.sql_text,
        record.notes,
        record.source_file,
        record.source_sheet,
        record.version,
        to_china_excel_datetime(
            record.created_at),
        to_china_excel_datetime(
            record.updated_at),
    ]


def _style_header(worksheet) -> None:
    fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )


def _set_widths(worksheet) -> None:
    widths = {
        "A": 18,
        "B": 12,
        "C": 16,
        "D": 14,
        "E": 18,
        "F": 35,
        "G": 10,
        "H": 16,
        "I": 14,
        "J": 40,
        "K": 16,
        "L": 16,
        "M": 35,
        "N": 35,
        "O": 35,
        "P": 35,
        "Q": 100,
        "R": 40,
        "S": 30,
        "T": 20,
        "U": 12,
        "V": 26,
        "W": 26,
    }

    for column, width in widths.items():
        worksheet.column_dimensions[
            column
        ].width = width

def format_datetime_columns(
    worksheet,
) -> None:
    """
    根据表头自动设置创建时间和更新时间列，
    不依赖固定的V列、W列。
    """

    datetime_headers = {
        "创建时间",
        "更新时间",
        "删除时间",
    }

    for header_cell in worksheet[1]:
        if (
            header_cell.value
            not in datetime_headers
        ):
            continue

        column_index = (
            header_cell.column
        )

        column_letter = (
            get_column_letter(
                column_index
            )
        )

        worksheet.column_dimensions[
            column_letter
        ].width = 21

        for row_index in range(
            2,
            worksheet.max_row + 1,
        ):
            cell = worksheet.cell(
                row=row_index,
                column=column_index,
            )

            cell.number_format = (
                EXCEL_DATETIME_FORMAT
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=False,
            )


def format_summary_time(
    worksheet,
) -> None:
    """
    设置导出说明工作表中的导出时间格式。
    """

    for row in worksheet.iter_rows():
        if row[0].value != "导出时间":
            continue

        time_cell = row[1]

        time_cell.number_format = (
            EXCEL_DATETIME_FORMAT
        )

        time_cell.alignment = Alignment(
            vertical="center",
            wrap_text=False,
        )

        worksheet.column_dimensions[
            get_column_letter(
                time_cell.column
            )
        ].width = 24

        break

def align_all_cells_left(
    worksheet,
) -> None:
    """
    将工作表中所有单元格设置为左对齐。

    使用copy保留原来的：
    1. 垂直对齐方式；
    2. 自动换行设置；
    3. 文本旋转；
    4. 缩进等其他样式。
    """

    for row in worksheet.iter_rows():
        for cell in row:
            alignment = copy(
                cell.alignment
            )

            alignment.horizontal = (
                "left"
            )

            cell.alignment = alignment

class ExportService:
    def __init__(self) -> None:
        self.audit_logs = (
            AuditLogRepository()
        )

    def export_sql_records(
        self,
        current_user: AuthenticatedUser,
    ) -> ExportResult:
        require_role(
            current_user,
            "viewer",
            "editor",
            "admin",
        )

        # 数据库和审计日志继续保存UTC时间
        exported_at_utc = utc_now()

        # Excel显示和文件名使用中国时间
        exported_at_china = (
            to_china_excel_datetime(
                exported_at_utc
            )
        )

        if exported_at_china is None:
            exported_at_china = (
                current_china_time()
            )

        with session_scope() as session:
            records = list(
                session.scalars(
                    select(SQLRecord)
                    .where(
                        SQLRecord.deleted_at
                        .is_(None)
                    )
                    .order_by(
                        SQLRecord.sql_id.asc()
                    )
                ).all()
            )

            workbook = Workbook()

            # =================================================
            # SQL详细语句工作表
            # =================================================

            worksheet = workbook.active
            worksheet.title = "SQL详细语句"

            worksheet.append(
                EXPORT_HEADERS
            )

            for record in records:
                worksheet.append(
                    _record_to_row(record)
                )

            _style_header(
                worksheet
            )

            _set_widths(
                worksheet
            )

            worksheet.freeze_panes = "A2"

            worksheet.auto_filter.ref = (
                worksheet.dimensions
            )

            # 隐藏数据库内部record_id
            worksheet.column_dimensions[
                "A"
            ].hidden = True

            # 先设置普通数据单元格样式
            for row in worksheet.iter_rows(
                min_row=2
            ):
                for cell in row:
                    cell.alignment = Alignment(
                        vertical="top",
                        wrap_text=True,
                    )

                worksheet.row_dimensions[
                    row[0].row
                ].height = 60

            # 必须放在普通样式设置之后，
            # 否则时间列的不换行设置会被覆盖
            format_datetime_columns(
                worksheet
            )
            align_all_cells_left(
                worksheet
            )

            # =================================================
            # 导出说明工作表
            # =================================================

            explanation = (
                workbook.create_sheet(
                    "导出说明",
                    0,
                )
            )

            explanation.append(
                [
                    "项目",
                    "内容",
                ]
            )

            explanation.append(
                [
                    "有效SQL数量",
                    len(records),
                ]
            )

            explanation.append(
                [
                    "导出用户",
                    current_user.username,
                ]
            )

            explanation.append(
                [
                    "导出时间",
                    exported_at_china,
                ]
            )

            explanation.append(
                [
                    "SQL_ID说明",
                    (
                        "SQL_ID是当前数据库中的"
                        "动态连续序号。"
                    ),
                ]
            )

            explanation.append(
                [
                    "记录ID说明",
                    (
                        "记录ID是系统内部稳定身份，"
                        "请勿删除或修改隐藏的A列。"
                    ),
                ]
            )

            _style_header(
                explanation
            )

            explanation.column_dimensions[
                "A"
            ].width = 24

            explanation.column_dimensions[
                "B"
            ].width = 90

            # 普通说明内容允许自动换行
            for row in explanation.iter_rows(
                min_row=2
            ):
                for cell in row:
                    cell.alignment = Alignment(
                        vertical="top",
                        wrap_text=True,
                    )

            # 必须放在普通说明样式之后
            format_summary_time(
                explanation
            )
            align_all_cells_left(
                explanation
            )
            # =================================================
            # 保存到内存
            # =================================================

            output = BytesIO()

            workbook.save(
                output
            )

            data = output.getvalue()

            output.close()
            workbook.close()

            # 审计日志保存UTC时间，这是正确的
            self.audit_logs.create(
                session,
                user_id=current_user.id,
                action="EXPORT",
                record_id=None,
                old_value=None,
                new_value=json.dumps(
                    {
                        "record_count": len(
                            records
                        ),
                        "exported_at": (
                            exported_at_utc
                            .isoformat()
                        ),
                        "display_timezone": (
                            "Asia/Shanghai"
                        ),
                    },
                    ensure_ascii=False,
                ),
            )

        # 文件名改为中国时间
        return ExportResult(
            data=data,
            file_name=(
                "SQL知识库_"
                + exported_at_china.strftime(
                    "%Y%m%d_%H%M%S"
                )
                + ".xlsx"
            ),
            record_count=len(records),
        )
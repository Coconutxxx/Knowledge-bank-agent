"""将常见文件统一读取为带来源位置的文本段。"""

from __future__ import annotations
import re
import csv
from dataclasses import dataclass
from pathlib import Path

from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader


@dataclass(frozen=True)
class LoadedSection:
    text: str
    source: str
    location: str


TEXT_EXTENSIONS = {".txt", ".md", ".sql", ".py", ".json", ".yaml", ".yml"}
SUPPORTED_EXTENSIONS = TEXT_EXTENSIONS | {".pdf", ".docx", ".csv", ".xlsx"}


def _read_text_with_fallback(path: Path) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    raise UnicodeDecodeError(
        "unknown", b"", 0, 1, f"无法识别文件编码：{path.name}"
    )


def _load_pdf(path: Path) -> list[LoadedSection]:
    reader = PdfReader(str(path))
    sections = []
    for page_number, page in enumerate(reader.pages, start=1):
        text = (page.extract_text() or "").strip()
        if text:
            sections.append(
                LoadedSection(text=text, source=path.name, location=f"第{page_number}页")
            )
    return sections


def _clean_excel_value(value) -> str:
    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip()

def _normalize_header(value) -> str:
    text = _clean_excel_value(value).lower()

    return re.sub(
        r"[\s_\-—:：()（）]+",
        "",
        text,
    )


def _find_header_column(
    row,
    column_type: str,
) -> int | None:
    normalized = [
        _normalize_header(cell)
        for cell in row
    ]

    if column_type == "sql_id":
        # SQL_ID、SQL ID、sql编号等写法
        for index, value in enumerate(normalized):
            if (
                value in {"sqlid", "sql编号", "sql序号"}
                or "sqlid" in value
            ):
                return index

    elif column_type == "function":
        # 优先寻找“功能主题”
        for index, value in enumerate(normalized):
            if value in {
                "功能主题",
                "功能名称",
                "sql功能",
                "查询功能",
            }:
                return index

        # 再匹配其他包含“功能”的列
        for index, value in enumerate(normalized):
            if (
                "功能" in value
                and "分类" not in value
                and "类型" not in value
            ):
                return index

    elif column_type == "sql_text":
        # 排除SQL_ID，只匹配SQL内容列
        preferred_names = {
            "sql正文",
            "sql语句",
            "sql详细语句",
            "标准sql",
            "hivesql",
            "sql内容",
        }

        for index, value in enumerate(normalized):
            if value in preferred_names:
                return index

        for index, value in enumerate(normalized):
            if (
                "sql" in value
                and "id" not in value
                and (
                    "正文" in value
                    or "语句" in value
                    or "内容" in value
                )
            ):
                return index

    return None


def _load_xlsx(path: Path) -> list[LoadedSection]:
    workbook = load_workbook(
        path,
        read_only=True,
        data_only=True,
    )

    sections = []

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            continue

        header_index = None
        sql_id_column = None
        function_column = None
        sql_text_column = None

        # 在前50行中寻找真正的标题行
        for row_index, row in enumerate(rows[:50]):
            current_sql_id_column = _find_header_column(
                row,
                "sql_id",
            )
            current_function_column = _find_header_column(
                row,
                "function",
            )
            current_sql_text_column = _find_header_column(
                row,
                "sql_text",
            )

            if (
                current_sql_id_column is not None
                and current_function_column is not None
                and current_sql_text_column is not None
            ):
                header_index = row_index
                sql_id_column = current_sql_id_column
                function_column = current_function_column
                sql_text_column = current_sql_text_column
                break

        # 不属于SQL功能明细的工作表不导入
        if header_index is None:
            continue

        headers = [
            _clean_excel_value(cell)
            or f"第{column + 1}列"
            for column, cell in enumerate(rows[header_index])
        ]

        useful_header_names = {
            "业务域",
            "步骤",
            "功能类型",
            "语句类型",
            "来源行号",
            "涉及来源表",
            "表资料完整性",
            "字段资料完整性",
            "缺失占位表",
            "未登记字段",
            "字段归属不明",
            "未解释字段",
            "缺少释义字段",
            "原始标题注释",
            "备注",
        }

        for excel_row_number, row in enumerate(
            rows[header_index + 1:],
            start=header_index + 2,
        ):
            values = [
                _clean_excel_value(cell)
                for cell in row
            ]

            def value_at(column_index: int | None) -> str:
                if column_index is None:
                    return ""

                if column_index >= len(values):
                    return ""

                return values[column_index]

            sql_id = value_at(sql_id_column)
            function_name = value_at(function_column)
            sql_text = value_at(sql_text_column)

            # SQL_ID、功能主题、SQL正文全部为空才跳过
            if not sql_id and not function_name and not sql_text:
                continue

            parts = [
                f"SQL_ID: {sql_id or excel_row_number}",
                f"功能主题: {function_name}",
            ]

            # 保存可用于回答的业务信息
            for column_index, header in enumerate(headers):
                if column_index in {
                    sql_id_column,
                    function_column,
                    sql_text_column,
                }:
                    continue

                normalized_header = _normalize_header(header)

                if normalized_header not in useful_header_names:
                    continue

                value = value_at(column_index)

                if value:
                    parts.append(
                        f"{header}: {value}"
                    )

            if sql_text:
                parts.extend(
                    [
                        "SQL正文开始",
                        sql_text,
                        "SQL正文结束",
                    ]
                )

            sections.append(
                LoadedSection(
                    text="\n".join(parts),
                    source=path.name,
                    location=(
                        f"工作表：{sheet.title}"
                        f"｜SQL_ID：{sql_id or excel_row_number}"
                    ),
                )
            )

    return sections

def _load_csv(path: Path) -> list[LoadedSection]:
    raw = _read_text_with_fallback(path)
    rows = []
    for row in csv.reader(raw.splitlines()):
        rows.append(" | ".join(cell.strip() for cell in row))
    text = "\n".join(rows).strip()
    return [LoadedSection(text=text, source=path.name, location="全文")] if text else []


def load_document(path: str | Path) -> list[LoadedSection]:
    file_path = Path(path)
    suffix = file_path.suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        raise ValueError(
            f"暂不支持 {suffix or '无扩展名'} 文件。"
            f"支持：{', '.join(sorted(SUPPORTED_EXTENSIONS))}"
        )
    if suffix in TEXT_EXTENSIONS:
        text = _read_text_with_fallback(file_path).strip()
        return [
            LoadedSection(text=text, source=file_path.name, location="全文")
        ] if text else []
    if suffix == ".pdf":
        return _load_pdf(file_path)
    if suffix == ".docx":
        return _load_docx(file_path)
    if suffix == ".csv":
        return _load_csv(file_path)
    return _load_xlsx(file_path)


def iter_supported_files(path: str | Path) -> list[Path]:
    target = Path(path)
    if target.is_file():
        return [target] if target.suffix.lower() in SUPPORTED_EXTENSIONS else []
    return sorted(
        file_path
        for file_path in target.rglob("*")
        if file_path.is_file() and file_path.suffix.lower() in SUPPORTED_EXTENSIONS
    )

